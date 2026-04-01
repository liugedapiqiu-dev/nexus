"""重新生成考勤异常汇总：从每日明细重建"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

SRC = '/home/user/Desktop/2026_03_考勤分析.xlsx'

wb = load_workbook(SRC)

# ── 从每日明细读取原始工作日数据 ─────────────────────────────────────────
ws_detail = wb['每日明细']
# 第3行是表头
headers = [ws_detail.cell(3, c).value for c in range(1, 13)]
print("每日明细表头:", headers)

rows_data = []
for r in range(4, ws_detail.max_row + 1):
    row = [ws_detail.cell(r, c).value for c in range(1, 13)]
    if any(v is not None for v in row):
        rows_data.append(row)

print(f"每日明细数据行: {len(rows_data)}")

# ── 过滤出有异常的记录 ───────────────────────────────────────────────────
abnormal_raw = [r for r in rows_data if
    (r[7] or 0) > 0 or (r[8] or 0) > 0 or (r[9] or 0) > 0 or r[10] == '是']

print(f"异常记录: {len(abnormal_raw)} 条")

# ── 按人聚合 ──────────────────────────────────────────────────────────────
grouped = defaultdict(lambda: {'rows': []})
for r in abnormal_raw:
    name = r[1]
    grouped[name]['rows'].append(r)

# 合并每人所有异常，生成汇总备注
# 保留第一条记录（基本信息），合并备注
final_rows = []
for name, data in grouped.items():
    first = data['rows'][0]
    all_parts = []
    for r in data['rows']:
        day = r[3].split('-')[-1] if r[3] else ''
        late  = r[7]  or 0
        early = r[8]  or 0
        ot    = r[9]  or 0
        absent= r[10]
        if late > 0:
            all_parts.append(f"{day}号迟到{late}分钟")
        if early > 0:
            all_parts.append(f"{day}号早退{early}分钟")
        if ot > 0:
            all_parts.append(f"{day}号加班{ot}分钟")
        if absent == '是':
            all_parts.append(f"{day}号缺勤")
    merged_note = '；'.join(all_parts)
    final_rows.append((*first[:11], merged_note))

print(f"聚合后人数: {len(final_rows)}")

# ── 重建考勤异常汇总 ───────────────────────────────────────────────────────
ws = wb['考勤异常汇总']
ws.delete_rows(1, ws.max_row)

HDR_FILL   = PatternFill("solid", fgColor="4472C4")
HDR_FONT   = Font(color="FFFFFF", bold=True)
LATE_FILL  = PatternFill("solid", fgColor="FFC7CE")
OT_FILL    = PatternFill("solid", fgColor="C6EFCE")
ABSENT_FILL= PatternFill("solid", fgColor="FFEB9C")
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center')

new_headers = ['工号','姓名','部门','日期','星期','签到时间','签退时间',
               '迟到(分)','早退(分)','加班(分)','缺勤','迟到/早退/加班/缺勤汇总备注']

ws.cell(1, 1, "2026年3月 考勤异常汇总").font = Font(bold=True, size=15)
ws.cell(2, 1, "统计时间: 2026-04-01  |  规则: 9:01起迟到 / 18:00前早退 / 22:00后加班 / 无记录缺勤  |  仅统计工作日(周一~五)").font = Font(italic=True, size=9, color="666666")

for c, h in enumerate(new_headers, 1):
    cell = ws.cell(3, c, h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = CENTER; cell.border = BORDER

for ri, row in enumerate(final_rows, 4):
    for ci, val in enumerate(row, 1):
        cell = ws.cell(ri, ci, val)
        cell.border = BORDER; cell.alignment = CENTER
        late  = row[7]  if row[7]  else 0
        early = row[8]  if row[8]  else 0
        ot    = row[9]  if row[9]  else 0
        absent= row[10] if row[10] else '否'
        if ci == 8  and late > 0:  cell.fill = LATE_FILL
        if ci == 9  and early > 0: cell.fill = LATE_FILL
        if ci == 10 and ot > 0:    cell.fill = OT_FILL
        if ci == 11 and absent == '是': cell.fill = ABSENT_FILL

col_widths = {1:12, 2:10, 3:20, 4:14, 5:6, 6:10, 7:10, 8:10, 9:10, 10:10, 11:8, 12:70}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

wb.save(SRC)
print(f"\n✅ 已保存: {SRC}")

# 预览
print("\n预览（前5人）:")
for r in final_rows[:5]:
    print(f"  {r[1]} | 迟到{r[7]}分 | 早退{r[8]}分 | 加班{r[9]}分 | 缺勤{r[10]} | {r[11]}")
