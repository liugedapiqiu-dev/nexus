"""
2026年3月考勤分析脚本
规则：
  上班签到 9:00（9:01起算迟到）
  下班签退 18:00（18:00前算早退）
  签退 > 22:00  → 加班
  无签到也无签退 → 缺勤
  仅统计工作日（周一~五）
"""

import re, calendar
from bs4 import BeautifulSoup
from datetime import datetime, time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
SRC = '/home/user/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/t7688103_6354/msg/file/2026-04/1-31.xls'
OUT = '/home/user/Desktop/2026_03_考勤分析.xlsx'

WORK_START    = time(9, 0)    # 9:00 整不算迟到
WORK_END      = time(18, 0)   # 18:00 整不算早退
OT_THRESHOLD  = time(22, 0)   # 22:00 后算加班

YEAR, MONTH = 2026, 3

# ─────────────────────────────────────────────────────────────────────────────
# 1. 解析 HTML
# ─────────────────────────────────────────────────────────────────────────────
with open(SRC, 'r', encoding='utf-8', errors='ignore') as f:
    html = f.read()

soup = BeautifulSoup(html, 'lxml')
table = soup.find('table')
rows  = table.find_all('tr')

# 找所有员工元数据行（10列，col0='编号'）
emp_start_rows = []
for i, r in enumerate(rows):
    cells = r.find_all(['td','th'])
    if len(cells) == 10 and cells[0].get_text(strip=True) == '编号':
        emp_start_rows.append(i)

print(f"共 {len(emp_start_rows)} 名员工")

employees = []
for idx in emp_start_rows:
    meta_cells = rows[idx].find_all(['td','th'])
    emp_id    = meta_cells[1].get_text(strip=True)
    emp_name  = meta_cells[3].get_text(strip=True)
    dept      = meta_cells[5].get_text(strip=True)
    join_date = meta_cells[7].get_text(strip=True)
    duty      = meta_cells[9].get_text(strip=True)

    date_row_idx = idx + 1
    sign_in_idx  = idx + 2
    sign_out_idx = idx + 3

    date_cells = rows[date_row_idx].find_all(['td','th'])
    sin_cells  = rows[sign_in_idx].find_all(['td','th'])
    sout_cells = rows[sign_out_idx].find_all(['td','th'])

    # col0 = '日期'/'签到1'/'签退1'，col1=day1, col2=day2, ... col31=day31
    # days[j] 对应第 j+1 天
    days = [date_cells[j].get_text(strip=True) for j in range(1, 32)]
    sin  = [sin_cells[j].get_text(strip=True)  for j in range(1, 32)]
    sout = [sout_cells[j].get_text(strip=True) for j in range(1, 32)]

    employees.append({
        'emp_id': emp_id, 'emp_name': emp_name, 'dept': dept,
        'join_date': join_date, 'duty': duty,
        'days': days, 'sign_in': sin, 'sign_out': sout,
        'html_row': idx,
    })

# ─────────────────────────────────────────────────────────────────────────────
# 2. 工作日 + 解析函数
# ─────────────────────────────────────────────────────────────────────────────
non_workdays = set()
for d in range(1, calendar.monthrange(YEAR, MONTH)[1] + 1):
    wd = calendar.weekday(YEAR, MONTH, d)
    if wd >= 5:           # 周六日
        non_workdays.add(d)

WEEKDAY_NAMES = ['一','二','三','四','五','六','日']

def parse_time(s):
    s = str(s).strip()
    if s in ('-', '', 'None', 'nan'): return None
    m = re.match(r'(\d{1,2}):(\d{2}):(\d{2})', s)
    if m: return time(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None

def tdelta_min(t_later, t_earlier):
    if t_later is None or t_earlier is None: return 0
    from datetime import timedelta
    diff = datetime.combine(datetime.today(), t_later) - datetime.combine(datetime.today(), t_earlier)
    return int(diff.total_seconds() / 60)

# ─────────────────────────────────────────────────────────────────────────────
# 3. 逐日分析
# ─────────────────────────────────────────────────────────────────────────────
all_rows  = []   # 全部工作日记录（含正常）
proc_rows = []   # 过程数据

for emp in employees:
    for day_idx in range(31):   # 3月有31天
        day_num = day_idx + 1
        wd = calendar.weekday(YEAR, MONTH, day_num)
        if wd >= 5: continue       # 排除周末

        day_str = emp['days'][day_idx]
        if not re.match(r'^\d+$', str(day_str)): continue

        si_val = emp['sign_in'][day_idx]
        so_val = emp['sign_out'][day_idx]
        si_t   = parse_time(si_val)
        so_t   = parse_time(so_val)

        late_min  = 0
        early_min = 0
        ot_min    = 0
        absent    = False
        notes     = []

        if si_t is None and so_t is None:
            absent = True
            notes.append('缺勤')
        else:
            if si_t is not None and si_t > WORK_START:
                late_min = tdelta_min(si_t, WORK_START)
                notes.append(f'迟到{late_min}分钟')
            if so_t is not None and so_t < WORK_END:
                early_min = tdelta_min(WORK_END, so_t)
                notes.append(f'早退{early_min}分钟')
            if so_t is not None and so_t > OT_THRESHOLD:
                ot_min = tdelta_min(so_t, OT_THRESHOLD)
                notes.append(f'加班{ot_min}分钟')

        status = '; '.join(notes) if notes else '正常'
        wd_name = WEEKDAY_NAMES[calendar.weekday(YEAR, MONTH, day_num)] if calendar.weekday(YEAR, MONTH, day_num) <= 4 else '周末'

        all_rows.append({
            '工号':     emp['emp_id'],
            '姓名':     emp['emp_name'],
            '部门':     emp['dept'],
            '日期':     f'2026-03-{day_num:02d}',
            '星期':     wd_name,
            '签到时间': si_val if si_val != '-' else '无记录',
            '签退时间': so_val if so_val != '-' else '无记录',
            '迟到(分)': late_min,
            '早退(分)': early_min,
            '加班(分)': ot_min,
            '缺勤':    '是' if absent else '否',
            '备注':     status,
        })

        proc_rows.append({
            '工号':     emp['emp_id'],
            '姓名':     emp['emp_name'],
            '部门':     emp['dept'],
            '日期':     f'2026-03-{day_num:02d}',
            '原始签到': si_val,
            '原始签退': so_val,
            '签到解析': str(si_t),
            '签退解析': str(so_t),
            '迟到(分)': late_min,
            '早退(分)': early_min,
            '加班(分)': ot_min,
            '缺勤':    '是' if absent else '否',
            '备注':     status,
        })

df_all = pd.DataFrame(all_rows)
print(f"工作日记录共 {len(df_all)} 条")

# ─────────────────────────────────────────────────────────────────────────────
# 4. 生成 XLSX
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# 样式颜色
C_HDR    = "4472C4"   # 蓝色表头
C_LATE   = "FFC7CE"   # 红色-迟到
C_EARLY  = "FFC7CE"   # 红色-早退
C_OT     = "C6EFCE"   # 绿色-加班
C_ABSENT = "FFEB9C"   # 黄色-缺勤

def mkfill(hex_color):  return PatternFill("solid", fgColor=hex_color)
def mkborder():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)
def mkalign(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

HDR_FILL   = mkfill(C_HDR)
HDR_FONT   = Font(color="FFFFFF", bold=True)
BORDER     = mkborder()

def write_ws(ws, title, subtitle, headers, data, extra_rows=0, highlight_cols=None):
    """写一个工作表"""
    ws.cell(1, 1, title).font = Font(bold=True, size=15)
    if subtitle:
        ws.cell(2, 1, subtitle).font = Font(italic=True, size=9, color="666666")

    start = 3 + extra_rows
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start, c, h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = mkalign()
        cell.border    = BORDER

    for r, row in enumerate(data, start+1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border    = BORDER
            cell.alignment = mkalign(wrap=True)
            # 高亮
            if highlight_cols:
                for col_idx, color in highlight_cols:
                    if c == col_idx and val not in (0, '0', '', None, '否', '正常'):
                        cell.fill = mkfill(color)

    # 列宽
    for col in range(1, len(headers)+1):
        cl = get_column_letter(col)
        ws.column_dimensions[cl].width = 14
    # 个别列加宽
    for cl, w in [('C', 20), ('D', 12), ('E', 14), ('F', 10), ('G', 10)]:
        ws.column_dimensions[cl].width = w

# ── Sheet1: 考勤异常汇总 ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "考勤异常汇总"

abn = df_all[(df_all['迟到(分)']>0) | (df_all['早退(分)']>0) | (df_all['加班(分)']>0) | (df_all['缺勤']=='是')].copy()
abn = abn.sort_values(['姓名','日期'])

headers1 = ['工号','姓名','部门','日期','星期','签到时间','签退时间','迟到(分)','早退(分)','加班(分)','缺勤','备注']
write_ws(ws1,
         "2026年3月 考勤异常汇总",
         f"统计时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  规则: 9:01起迟到 / 18:00前早退 / 22:00后加班 / 无记录缺勤  |  仅统计工作日(周一~五)",
         headers1,
         abn[['工号','姓名','部门','日期','星期','签到时间','签退时间','迟到(分)','早退(分)','加班(分)','缺勤','备注']].values.tolist(),
         highlight_cols=[(8, C_LATE), (9, C_EARLY), (10, C_OT), (11, C_ABSENT)])

# ── Sheet2: 每日明细 ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("每日明细")
detail = df_all.sort_values(['姓名','日期'])
write_ws(ws2,
         "2026年3月 每日考勤明细（仅工作日）",
         f"共计 {len(detail)} 条工作日记录  |  红色=迟到/早退  绿色=加班  黄色=缺勤",
         headers1,
         detail[['工号','姓名','部门','日期','星期','签到时间','签退时间','迟到(分)','早退(分)','加班(分)','缺勤','备注']].values.tolist(),
         highlight_cols=[(8, C_LATE), (9, C_EARLY), (10, C_OT), (11, C_ABSENT)])

# ── Sheet3: 人员考勤透视 ─────────────────────────────────────────────────
ws3 = wb.create_sheet("人员考勤透视")

from collections import defaultdict
stats = defaultdict(lambda: {'late':[], 'early':[], 'ot':[], 'absent':0})
for _, r in df_all.iterrows():
    k = (r['工号'], r['姓名'], r['部门'])
    if r['迟到(分)'] > 0:
        stats[k]['late'].append((r['日期'], r['迟到(分)']))
    if r['早退(分)'] > 0:
        stats[k]['early'].append((r['日期'], r['早退(分)']))
    if r['加班(分)'] > 0:
        stats[k]['ot'].append((r['日期'], r['加班(分)']))
    if r['缺勤'] == '是':
        stats[k]['absent'] += 1

ph = ['工号','姓名','部门','迟到次数','迟到明细','早退次数','早退明细','加班次数','加班明细','缺勤天数']
pdata = []
for (eid, ename, edept), s in sorted(stats.items()):
    pdata.append([
        eid, ename, edept,
        len(s['late']),
        '; '.join(f"{d}({m}分)" for d,m in s['late']) or '-',
        len(s['early']),
        '; '.join(f"{d}({m}分)" for d,m in s['early']) or '-',
        len(s['ot']),
        '; '.join(f"{d}({m}分)" for d,m in s['ot']) or '-',
        s['absent'],
    ])

write_ws(ws3,
         "2026年3月 人员考勤透视",
         "迟到/早退/加班：次数+明细(含日期和分钟)；缺勤：天数",
         ph, pdata,
         extra_rows=1,
         highlight_cols=[(4, C_LATE), (6, C_EARLY), (8, C_OT), (10, C_ABSENT)])

# 加宽明细列
ws3.column_dimensions['E'].width = 50
ws3.column_dimensions['G'].width = 50
ws3.column_dimensions['I'].width = 50

# ── Sheet4: 原始解析过程 ─────────────────────────────────────────────────
ws4 = wb.create_sheet("原始解析过程")
ph4 = ['工号','姓名','部门','日期','原始签到','原始签退','签到解析','签退解析','迟到(分)','早退(分)','加班(分)','缺勤','备注']
pdata4 = [[r['工号'],r['姓名'],r['部门'],r['日期'],r['原始签到'],r['原始签退'],
           r['签到解析'],r['签退解析'],r['迟到(分)'],r['早退(分)'],r['加班(分)'],r['缺勤'],r['备注']]
          for r in proc_rows]

write_ws(ws4,
         "2026年3月 原始数据解析过程",
         f"数据源: {SRC}  |  解析规则: 9:01起迟到;18:00前早退;22:00后加班;无签到签退=缺勤;仅统计工作日",
         ph4, pdata4,
         extra_rows=1,
         highlight_cols=[(9, C_LATE), (10, C_EARLY), (11, C_OT), (12, C_ABSENT)])

# ── Sheet5: 考勤规则说明 ─────────────────────────────────────────────────
ws5 = wb.create_sheet("考勤规则说明")
rules = [
    ("项目", "说明"),
    ("上班签到", "9:00整不算迟到；9:01起算迟到"),
    ("下班签退", "18:00整不算早退；18:00前算早退"),
    ("加班判定", "签退时间在22:00之后，记为加班"),
    ("缺勤判定", "当天无签到记录且无签退记录，判定为缺勤"),
    ("工作日", "周一至周五；3月无法定假日，仅排除周末"),
    ("周末", "周六、周日数据不纳入统计"),
    ("迟到计算", "迟到分钟 = 实际签到时间 - 09:00"),
    ("早退计算", "早退分钟 = 18:00 - 实际签退时间"),
    ("加班计算", "加班分钟 = 实际签退时间 - 22:00"),
    ("颜色说明", "红色=迟到/早退；绿色=加班；黄色=缺勤"),
    ("", ""),
    ("输出表单", ""),
    ("考勤异常汇总", "仅显示有迟到、早退、加班或缺勤的记录"),
    ("每日明细", "全部工作日的完整考勤记录，含原始打卡时间"),
    ("人员考勤透视", "按人员汇总各项次数，并列出明细(日期/分钟)"),
    ("原始解析过程", "数据解析中间步骤，含原始值、解析结果、判定依据"),
]
for r, (k, v) in enumerate(rules, 1):
    c1 = ws5.cell(r, 1, k)
    c2 = ws5.cell(r, 2, v)
    if k in ("项目","输出表单"):
        c1.font = Font(bold=True)
    if k in ("颜色说明",):
        c2.font = Font(color="666666")
ws5.column_dimensions['A'].width = 18
ws5.column_dimensions['B'].width = 65

# ─────────────────────────────────────────────────────────────────────────────
# 保存
wb.save(OUT)
print(f"\n✅ 已保存: {OUT}")
print(f"   员工总数: {len(employees)}")
print(f"   工作日记录: {len(df_all)} 条")
abn_count = len(df_all[(df_all['迟到(分)']>0)|(df_all['早退(分)']>0)|(df_all['加班(分)']>0)|(df_all['缺勤']=='是')])
print(f"   异常记录: {abn_count} 条")
print(f"   异常率: {abn_count/len(df_all)*100:.1f}%")
