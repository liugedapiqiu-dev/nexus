"""给考勤异常汇总增加一行一人备注"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

SRC = '/home/user/Desktop/2026_03_考勤分析.xlsx'

wb = load_workbook(SRC)
ws = wb['考勤异常汇总']

# ── 1. 读取现有数据（从第4行开始，第3行是表头）─────────────────────────────
# 读取表头（第3行，index=2）
headers = [ws.cell(3, c).value for c in range(1, ws.max_column+1)]
print("表头:", headers)

# 读取所有数据行
rows_data = []
for r in range(4, ws.max_row+1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    if any(v is not None for v in row):
        rows_data.append(row)

print(f"现有数据行数: {len(rows_data)}")

# ── 2. 按人聚合，生成备注 ──────────────────────────────────────────────────
# 格式：13号迟到3分钟，31号迟到4分钟
grouped = defaultdict(lambda: {'late': [], 'early': [], 'ot': [], 'absent': []})

for row in rows_data:
    name = row[1]   # 姓名
    date_str = row[3]  # 2026-03-DD
    day = date_str.split('-')[-1] if date_str else ''
    late  = row[7]  if row[7]  else 0  # 迟到(分)
    early = row[8]  if row[8]  else 0  # 早退(分)
    ot    = row[9]  if row[9]  else 0  # 加班(分)
    absent= row[10] if row[10] else '否'  # 缺勤

    if late > 0:
        grouped[name]['late'].append(f"{day}号迟到{late}分钟")
    if early > 0:
        grouped[name]['early'].append(f"{day}号早退{early}分钟")
    if ot > 0:
        grouped[name]['ot'].append(f"{day}号加班{ot}分钟")
    if absent == '是':
        grouped[name]['absent'].append(f"{day}号缺勤")

# 构建合并后的行
def build_note(parts):
    return '；'.join(parts)

merged_rows = []
for name, stats in grouped.items():
    all_parts = stats['late'] + stats['early'] + stats['ot'] + stats['absent']
    note = build_note(all_parts)
    # 取该人第一条记录作为基础
    first = next(r for r in rows_data if r[1] == name)
    merged_rows.append((*first[:11], note))  # 前11列原样，第12列新备注

print(f"聚合后行数: {len(merged_rows)}")
print("示例备注:")
for r in merged_rows[:5]:
    print(f"  {r[1]}: {r[11]}")

# ── 3. 重建考勤异常汇总 ─────────────────────────────────────────────────────
# 样式
HDR_FILL  = PatternFill("solid", fgColor="4472C4")
HDR_FONT  = Font(color="FFFFFF", bold=True)
LATE_FILL = PatternFill("solid", fgColor="FFC7CE")
OT_FILL   = PatternFill("solid", fgColor="C6EFCE")
ABSENT_FILL=PatternFill("solid", fgColor="FFEB9C")
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center')

# 原12列 → 保持12列，把旧"备注"列替换成汇总备注（覆盖旧备注）
new_headers = ['工号','姓名','部门','日期','星期','签到时间','签退时间','迟到(分)','早退(分)','加班(分)','缺勤','迟到/早退/加班/缺勤汇总备注']

# 清空并重建
ws.delete_rows(1, ws.max_row)

# 标题行
ws.cell(1, 1, "2026年3月 考勤异常汇总").font = Font(bold=True, size=15)
ws.cell(2, 1, "统计时间: 2026-04-01  |  规则: 9:01起迟到 / 18:00前早退 / 22:00后加班 / 无记录缺勤  |  仅统计工作日(周一~五)").font = Font(italic=True, size=9, color="666666")

# 表头（第3行）
for c, h in enumerate(new_headers, 1):
    cell = ws.cell(3, c, h)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = CENTER
    cell.border = BORDER

# 数据行
for ri, row in enumerate(merged_rows, 4):
    for ci, val in enumerate(row, 1):
        cell = ws.cell(ri, ci, val)
        cell.border = BORDER
        cell.alignment = CENTER
        # 高亮
        late  = row[7]  if row[7]  else 0
        early = row[8]  if row[8]  else 0
        ot    = row[9]  if row[9]  else 0
        absent= row[10] if row[10] else '否'
        if ci == 8  and late > 0:  cell.fill = LATE_FILL
        if ci == 9  and early > 0: cell.fill = LATE_FILL
        if ci == 10 and ot > 0:    cell.fill = OT_FILL
        if ci == 11 and absent == '是': cell.fill = ABSENT_FILL

# 列宽
col_widths = {1:12, 2:10, 3:20, 4:14, 5:6, 6:10, 7:10, 8:10, 9:10, 10:10, 11:8, 12:60}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

wb.save(SRC)
print(f"\n✅ 已更新: {SRC}")
print(f"   聚合后员工数: {len(merged_rows)}")
